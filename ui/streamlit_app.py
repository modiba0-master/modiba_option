"""
Streamlit UI - 모디바 옵션 가격 계산기
옵션별 가중치 방식: 변경가 = 대표 기준가 × 옵션가중치
"""

import asyncio
import sys
import os
import re
import io

import streamlit as st
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

from app.services import naver_api
from app.models.product_model import CalculatedOption

# ──────────────────────────────────────────────
# 페이지 설정
# ──────────────────────────────────────────────
st.set_page_config(
    page_title="모디바 옵션 가격 계산기",
    page_icon="🧮",
    layout="wide",
)

DEFAULT_PRODUCT_ID = os.getenv("DEFAULT_PRODUCT_ID", "6774969928")

# ──────────────────────────────────────────────
# CSS
# ──────────────────────────────────────────────
st.markdown("""
<style>
[data-testid="stMetricValue"] { font-size: 1.4rem; font-weight: 700; }
[data-testid="stMetricLabel"] { font-size: 0.8rem; color: #666; }
.block-container { padding-top: 1.2rem; padding-bottom: 2rem; }
div[data-testid="stDataFrame"] { border: 1px solid #e0e0e0; border-radius: 8px; }
.stAlert { border-radius: 8px; }
</style>
""", unsafe_allow_html=True)


# ──────────────────────────────────────────────
# 유틸리티
# ──────────────────────────────────────────────
def run_async(coro):
    try:
        loop = asyncio.get_event_loop()
        if loop.is_running():
            import concurrent.futures
            with concurrent.futures.ThreadPoolExecutor() as pool:
                return pool.submit(asyncio.run, coro).result()
        return loop.run_until_complete(coro)
    except RuntimeError:
        return asyncio.run(coro)


def suggest_weight(naver_actual_price: int, base_price: int) -> float:
    """
    네이버 원본 실판매가(= 초기할인가 + 옵션가)를 기준가로 나눠 가중치를 역산합니다.
    weight = naver_actual_price / base_price
    """
    if base_price <= 0 or naver_actual_price <= 0:
        return 1.0
    return round(naver_actual_price / base_price, 2)


def get_weight_key(option_id: str) -> str:
    return f"opt_w_{option_id}"


def build_option_excel(
    product_name: str,
    product_id: str,
    sale_price: int,
    discount_amount: int,
    naver_discounted: int,
    base_price: int,
    new_sale_price: int,
    discount_rate: int,
    delivery_fee: int,
    option_rows: list,   # list of dict (⑤ 테이블 행 데이터)
) -> bytes:
    """
    ⑤ 옵션별 가중치 설정 테이블 기준으로 엑셀을 생성합니다.
    option_rows: [{"옵션명", "재고", "기존옵션가", "기존실판매가", "가중치",
                   "변경계산가", "기준가대비", "기존단가", "변경단가", "순단가"}, ...]
    """
    # ── 스타일 상수
    NAVY   = PatternFill("solid", fgColor="1F4E79")
    GRAY   = PatternFill("solid", fgColor="D9E1F2")
    BLUE   = PatternFill("solid", fgColor="DEEAF1")
    WHITE  = Font(color="FFFFFF", bold=True, name="맑은 고딕", size=10)
    BOLD   = Font(bold=True, name="맑은 고딕", size=10)
    NORMAL = Font(name="맑은 고딕", size=10)
    CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
    RIGHT  = Alignment(horizontal="right", vertical="center")
    LEFT   = Alignment(horizontal="left", vertical="center", wrap_text=True)
    THIN   = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "옵션가격계산"
    ws.sheet_view.showGridLines = False

    row = 1

    def cell(r, c, val, font=NORMAL, fill=None, align=CENTER, border=THIN):
        _c = ws.cell(row=r, column=c, value=val)
        _c.font = font
        if fill: _c.fill = fill
        _c.alignment = align
        _c.border = border
        return _c

    # ── 1. 상품 헤더 (2행)
    ws.merge_cells(f"A{row}:D{row}")
    cell(row, 1, "상품 정보", font=WHITE, fill=NAVY, align=CENTER)
    ws.merge_cells(f"E{row}:J{row}")
    cell(row, 5, "기준가 설정", font=WHITE, fill=NAVY, align=CENTER)
    row += 1

    headers_info = ["상품ID", "상품명", "판매가", "기본할인", "네이버 할인가(기준)", "변경 기준가",
                    f"변경 판매가({discount_rate}% 할인)", "택배비", "", ""]
    values_info  = [product_id, product_name, f"{sale_price:,}원", f"{discount_amount:,}원",
                    f"{naver_discounted:,}원", f"{base_price:,}원",
                    f"{new_sale_price:,}원", f"{delivery_fee:,}원", "", ""]
    for col_i, (h, v) in enumerate(zip(headers_info[:8], values_info[:8]), start=1):
        cell(row, col_i, h, font=BOLD, fill=GRAY, align=CENTER)
        cell(row + 1, col_i, v, font=NORMAL, align=CENTER)
    row += 2
    ws.row_dimensions[row - 1].height = 18

    # 빈 구분행
    row += 1

    # ── 2. 옵션 테이블 헤더
    col_headers = [
        ("옵션명",        LEFT,  30),
        ("재고",          CENTER, 8),
        ("기존 옵션가",   CENTER, 12),
        ("기존 실판매가", CENTER, 13),
        ("가중치",        CENTER, 8),
        ("변경 계산가",   CENTER, 13),
        ("기준가 대비",   CENTER, 12),
        ("기존 단가",     CENTER, 13),
        ("변경 단가",     CENTER, 13),
        ("순단가\n(택배제외)", CENTER, 13),
    ]
    for col_i, (header, align, width) in enumerate(col_headers, start=1):
        cell(row, col_i, header, font=WHITE, fill=NAVY, align=CENTER)
        ws.column_dimensions[get_column_letter(col_i)].width = width
    ws.row_dimensions[row].height = 30
    row += 1

    # ── 3. 옵션 데이터 행
    GREEN = Font(color="1A7F37", bold=True, name="맑은 고딕", size=10)
    RED   = Font(color="D73A49", bold=True, name="맑은 고딕", size=10)

    for i, opt_row in enumerate(option_rows):
        fill_bg = PatternFill("solid", fgColor="EEF3FB") if i % 2 == 0 else None

        def data_cell(col, val, font=None, align=CENTER):
            _f = font or NORMAL
            _c = ws.cell(row=row, column=col, value=val)
            _c.font = _f
            if fill_bg: _c.fill = fill_bg
            _c.alignment = align
            _c.border = THIN
            return _c

        data_cell(1, opt_row["옵션명"], align=LEFT)
        data_cell(2, opt_row["재고"])
        data_cell(3, opt_row["기존옵션가"])
        data_cell(4, opt_row["기존실판매가"])
        data_cell(5, opt_row["가중치"])
        data_cell(6, opt_row["변경계산가"])

        # 기준가 대비: 양수=초록, 음수=빨강
        vs_val = opt_row["기준가대비"]
        vs_font = GREEN if vs_val >= 0 else RED
        vs_str = f"+{vs_val:,}원" if vs_val >= 0 else f"{vs_val:,}원"
        data_cell(7, vs_str, font=vs_font)

        data_cell(8, opt_row["기존단가"])
        data_cell(9, opt_row["변경단가"])
        data_cell(10, opt_row["순단가"])

        ws.row_dimensions[row].height = 16
        row += 1

    # ── 저장
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def parse_unit_price(option_name: str, calc_price: int) -> tuple:
    """
    옵션명에서 단위와 수량을 추출하여 단가를 계산합니다.
    반환: (라벨, 단가) 예: ("1kg당", 10600) 또는 ("—", None)

    인식 우선순위:
    1. 팩 패턴  : (1kgX3팩), (500gX6팩), (200gX5)  → 팩당 단가
    2. 총 kg    : "3kg", "10kg" 등                  → 1kg당 단가
    3. 총 g     : "500g", "200g" 등                 → 100g당 단가
    """
    # 1) 팩 패턴: XkgXN 또는 XgXN
    m = re.search(r'(\d+(?:\.\d+)?)(kg|g)[Xx×](\d+)', option_name, re.IGNORECASE)
    if m:
        size = float(m.group(1))
        unit = m.group(2).lower()
        count = int(m.group(3))
        if count > 0:
            label = f"{int(size)}kg당" if unit == "kg" else f"{int(size)}g당"
            return label, round(calc_price / count)

    # 2) 총 kg 중량: Nkg 숫자 중 가장 큰 값 기준 → 1kg당 단가
    kg_matches = re.findall(r'(\d+(?:\.\d+)?)\s*kg', option_name, re.IGNORECASE)
    if kg_matches:
        total_kg = max(float(x) for x in kg_matches)
        if total_kg > 0:
            return "1kg당", round(calc_price / total_kg)

    # 3) 총 g 중량: Ng 숫자 중 가장 큰 값 기준 → 100g당 단가
    g_matches = re.findall(r'(\d+(?:\.\d+)?)\s*g(?![a-zA-Z])', option_name, re.IGNORECASE)
    if g_matches:
        total_g = max(float(x) for x in g_matches)
        if total_g >= 100:
            return "100g당", round(calc_price * 100 / total_g)

    return "—", None


# ──────────────────────────────────────────────
# 사이드바
# ──────────────────────────────────────────────
def _get_server_ip() -> str:
    import urllib.request
    for url in ["https://api.ipify.org", "https://icanhazip.com"]:
        try:
            return urllib.request.urlopen(url, timeout=5).read().decode().strip()
        except Exception:
            continue
    return "조회 실패"


def _test_naver_api() -> dict:
    import base64, time
    import bcrypt, httpx
    cid = os.getenv("NAVER_COMMERCE_API_CLIENT_ID", "")
    csecret = os.getenv("NAVER_COMMERCE_API_CLIENT_SECRET", "")
    if not cid or not csecret:
        return {"ok": False, "msg": "환경변수 미설정"}
    ts = int(time.time() * 1000)
    hashed = bcrypt.hashpw(f"{cid}_{ts}".encode(), csecret.encode())
    sig = base64.b64encode(hashed).decode()
    try:
        r = httpx.post(
            "https://api.commerce.naver.com/external/v1/oauth2/token",
            data={"client_id": cid, "timestamp": ts, "client_secret_sign": sig,
                  "grant_type": "client_credentials", "type": "SELF"},
            headers={"Content-Type": "application/x-www-form-urlencoded"}, timeout=10,
        )
        d = r.json()
    except Exception as e:
        return {"ok": False, "msg": str(e)}
    if r.status_code != 200 or "access_token" not in d:
        return {"ok": False, "code": d.get("code", ""), "msg": d.get("message", "")[:100]}
    return {"ok": True, "token_prefix": d["access_token"][:20]}


with st.sidebar:
    cid_set = bool(os.getenv("NAVER_COMMERCE_API_CLIENT_ID", ""))
    if cid_set:
        st.success("✅ 네이버 API 연결됨")
    else:
        st.error("❌ API 자격증명 미설정")
    st.caption(f"기본 상품ID: **{DEFAULT_PRODUCT_ID}**")
    st.divider()
    with st.expander("🛠️ 개발자 도구", expanded=False):
        if st.button("🌐 서버 IP 확인", use_container_width=True):
            with st.spinner("조회 중..."):
                st.session_state["server_ip"] = _get_server_ip()
        if "server_ip" in st.session_state:
            st.code(st.session_state["server_ip"], language=None)
            st.caption("👆 네이버 API 화이트리스트 등록 IP")
        st.markdown("---")
        if st.button("🧪 API 연결 테스트", use_container_width=True):
            with st.spinner("테스트 중..."):
                st.session_state["api_test"] = _test_naver_api()
        if "api_test" in st.session_state:
            res = st.session_state["api_test"]
            if res.get("ok"):
                st.success("API 정상")
            else:
                st.error(f"{res.get('code','')} {res.get('msg','')}")
    st.divider()
    st.markdown(
        "<small>조회 + 계산 + 엑셀 출력 전용<br>가격 자동반영 기능 없음</small>",
        unsafe_allow_html=True,
    )


# ──────────────────────────────────────────────
# 헤더
# ──────────────────────────────────────────────
st.title("🧮 모디바 옵션 가격 계산기")
st.caption("대표 기준가를 설정하면 옵션별 가중치에 따라 변경가가 자동으로 계산됩니다.")
st.divider()


# ──────────────────────────────────────────────
# 세션 초기화
# ──────────────────────────────────────────────
for key, default in [
    ("search_results", []),
    ("selected_product", None),
    ("auto_loaded", False),
    ("last_query", ""),
    ("weights_initialized_for", ""),
    ("prev_base_price", 0),   # 기준가 변경 감지용
]:
    if key not in st.session_state:
        st.session_state[key] = default


# ──────────────────────────────────────────────
# 기본 상품 자동 로드 (최초 1회)
# ──────────────────────────────────────────────
if not st.session_state.auto_loaded:
    with st.spinner(f"상품 {DEFAULT_PRODUCT_ID} 불러오는 중..."):
        product = run_async(naver_api.get_product_detail(DEFAULT_PRODUCT_ID))
    if product:
        if not product.product_id:
            product = product.model_copy(update={"product_id": DEFAULT_PRODUCT_ID})
        st.session_state.selected_product = product
        st.session_state.search_results = [product]
    st.session_state.auto_loaded = True


# ──────────────────────────────────────────────
# ① 상품 검색
# ──────────────────────────────────────────────
st.subheader("① 상품 검색")
col_s, col_b = st.columns([5, 1])
with col_s:
    query = st.text_input("상품ID", placeholder="예: 6774969928",
                          label_visibility="collapsed")
with col_b:
    if st.button("🔍 검색", use_container_width=True):
        if query.strip():
            with st.spinner("검색 중..."):
                results = run_async(naver_api.search_products(query.strip()))
            for p in results:
                if not p.product_id and query.strip().isdigit():
                    p.product_id = query.strip()
            st.session_state.search_results = results
            st.session_state.selected_product = None
            st.session_state.last_query = query.strip()
            st.session_state.weights_initialized_for = ""
            if not results:
                st.warning("검색 결과가 없습니다.")


# ──────────────────────────────────────────────
# ② 상품 선택 (필터 테이블)
# ──────────────────────────────────────────────
if st.session_state.search_results:
    st.subheader("② 상품 선택")

    all_results = st.session_state.search_results

    # 실시간 필터 입력
    col_flt, col_cnt = st.columns([4, 1])
    with col_flt:
        filter_text = st.text_input(
            "상품 필터",
            placeholder="상품ID 또는 상품명 입력 시 실시간 필터...",
            label_visibility="collapsed",
            key="product_filter_input",
        )
    with col_cnt:
        st.caption(f"총 {len(all_results)}개 상품")

    # 필터 적용
    ft = filter_text.strip().lower()
    filtered_results = [
        p for p in all_results
        if not ft
        or ft in (p.product_id or "").lower()
        or ft in p.product_name.lower()
    ]

    if not filtered_results:
        st.warning("필터 조건에 맞는 상품이 없습니다.")
    else:
        # 테이블 표시용 DataFrame (상품ID + 상품명 모두 표시)
        df_select = pd.DataFrame([
            {
                "상품ID": p.product_id or st.session_state.last_query or DEFAULT_PRODUCT_ID,
                "상품명": p.product_name,
                "판매가": f"{p.sale_price:,}원",
                "할인가": f"{p.discounted_price:,}원",
                "옵션수": f"{len(p.options)}개",
            }
            for p in filtered_results
        ])

        tbl_height = min(250, (len(filtered_results) + 1) * 38 + 10)

        event = st.dataframe(
            df_select,
            use_container_width=True,
            hide_index=True,
            height=tbl_height,
            on_select="rerun",
            selection_mode="single-row",
        )

        # 행 선택 처리
        selected_rows = event.selection.rows if event.selection else []

        if selected_rows:
            new_product = filtered_results[selected_rows[0]]
            if (st.session_state.selected_product is None or
                    new_product.product_id != getattr(
                        st.session_state.selected_product, "product_id", None)):
                st.session_state.weights_initialized_for = ""
            st.session_state.selected_product = new_product
        elif len(filtered_results) == 1 and st.session_state.selected_product is None:
            # 결과 1개이면 자동 선택
            st.session_state.selected_product = filtered_results[0]

        # 현재 선택된 상품 표시
        if st.session_state.selected_product:
            sp = st.session_state.selected_product
            sp_pid = sp.product_id or DEFAULT_PRODUCT_ID
            st.caption(f"✅ 선택된 상품: **[{sp_pid}] {sp.product_name}**")


# ──────────────────────────────────────────────
# ③ 상품 상세 + ④ 기준가 설정
# ──────────────────────────────────────────────
if st.session_state.selected_product:
    product = st.session_state.selected_product
    display_id = product.product_id or DEFAULT_PRODUCT_ID
    st.divider()

    col_info, col_base = st.columns([2, 1])

    with col_info:
        st.subheader("③ 상품 상세")
        st.markdown(f"### {product.product_name}")
        st.caption(f"상품번호: {display_id} | 옵션 {len(product.options)}개")

        # 네이버 원본 가격 (고정 기준 — 변경 불가)
        st.caption("**[네이버 원본]** ← 기존 단가의 기준 (초기 할인가 + 옵션가)")
        m1, m2, m3 = st.columns(3)
        m1.metric("판매가", f"{product.sale_price:,}원")
        m2.metric("기본할인", f"▼ {product.discount_amount:,}원")
        m3.metric("할인가 (기준)", f"{product.discounted_price:,}원")

    with col_base:
        st.subheader("④ 대표 기준가 설정")

        # 할인율 설정
        discount_rate = st.number_input(
            "할인율 (%)",
            min_value=1,
            max_value=99,
            value=48,
            step=1,
            key="discount_rate_input",
            help="기준가 = 변경 판매가 × (1 - 할인율/100). 기본 48%",
        )

        # 기준가 직접 입력 — 명시적 key로 세션 고정 (값이 임의로 초기화되지 않음)
        default_base = (product.discounted_price if product.discounted_price > 0
                        else product.sale_price)

        # 상품이 바뀌면 세션의 base_price를 새 상품 기본값으로 리셋
        if st.session_state.weights_initialized_for != display_id:
            st.session_state["base_price_input"] = default_base

        base_price = st.number_input(
            "기준가 (원)",
            min_value=1,
            value=st.session_state.get("base_price_input", default_base),
            step=100,
            key="base_price_input",
            help="변경 계산가 = 기준가 × 옵션가중치. 변경 시 가중치가 자동으로 재산출됩니다.",
        )

        # 기준가 → 변경된 판매가 역산 (할인율 기반)
        rate = discount_rate / 100
        new_sale_price = round(base_price / (1 - rate)) if rate < 1 else base_price

        st.markdown("---")
        st.caption("**[기준가 기반 변경 가격]**")
        nb1, nb2 = st.columns(2)
        nb1.metric(
            "변경 기준가",
            f"{base_price:,}원",
            delta=f"{base_price - product.discounted_price:+,}원 vs 할인가",
            delta_color="off",
        )
        nb2.metric(
            f"변경 판매가 ({discount_rate}% 할인 기준)",
            f"{new_sale_price:,}원",
            delta=f"{new_sale_price - product.sale_price:+,}원 vs 네이버",
            delta_color="off",
        )
        st.caption(f"공식: {base_price:,} ÷ (1 - {discount_rate}%) = {new_sale_price:,}원")

    st.divider()

    # ──────────────────────────────────────────────
    # ⑤ 옵션별 가중치 설정 (실시간 자동 계산)
    # ──────────────────────────────────────────────
    st.subheader(f"⑤ 옵션별 가중치 설정 ({len(product.options)}개)")

    # 네이버 초기 할인가 (고정 기준 — 사용자 기준가 변경과 무관)
    naver_discounted = product.discounted_price if product.discounted_price > 0 else product.sale_price

    # 새 상품 로드 시에만 가중치 초기화 (기준가 변경 시 자동재산출 없음 — 수동 조정만)
    if st.session_state.weights_initialized_for != display_id:
        for opt in product.options:
            key = get_weight_key(opt.option_id)
            naver_actual = naver_discounted + opt.option_price
            # 초기 가중치 = 네이버 실판매가 / 네이버 할인가 (고정 비율)
            st.session_state[key] = suggest_weight(naver_actual, naver_discounted)
        st.session_state.weights_initialized_for = display_id

    col_hdr, col_delivery, col_reset = st.columns([3, 1.5, 1])
    with col_hdr:
        st.caption(
            "💡 가중치 = 기존실판매가 / 네이버할인가. 기준가 변경 시 가중치는 고정, "
            "변경단가만 재계산됩니다. 🔄 로 초기 비율 복원 가능."
        )
    with col_delivery:
        delivery_fee = st.number_input(
            "📦 택배비 (원)",
            min_value=0,
            value=3500,
            step=100,
            key="delivery_fee_input",
            help="변경 계산가에서 택배비를 차감한 순단가를 계산합니다.",
        )
    with col_reset:
        st.write("")
        if st.button("🔄 가중치 초기화", use_container_width=True):
            for opt in product.options:
                naver_actual = naver_discounted + opt.option_price
                st.session_state[get_weight_key(opt.option_id)] = suggest_weight(
                    naver_actual, naver_discounted   # 네이버 할인가 기준 비율
                )
            st.rerun()

    # 테이블 헤더 (10컬럼)
    # 옵션명 | 재고 | 기존옵션가 | 기존실판매가(new) | 가중치 | 변경계산가 | 기준가대비 | 기존단가 | 변경단가 | 순단가
    h0, h1, h2, h2b, h3, h4, h5, h6, h7, h8 = st.columns(
        [2.5, 0.7, 0.9, 1.0, 0.8, 1.0, 1.0, 1.0, 1.0, 1.0]
    )
    h0.markdown("**옵션명**")
    h1.markdown("**재고**")
    h2.markdown("**기존 옵션가**<br><small style='color:#888;font-weight:normal'>상대값</small>",
                unsafe_allow_html=True)
    h2b.markdown(f"**기존 실판매가**<br><small style='color:#888;font-weight:normal'>할인가({naver_discounted:,})+옵션가</small>",
                 unsafe_allow_html=True)
    h3.markdown("**가중치**")
    h4.markdown("**변경 계산가**<br><small style='color:#888;font-weight:normal'>가중치×기준가</small>",
                unsafe_allow_html=True)
    h5.markdown("**기준가 대비**")
    h6.markdown("**기존 단가**")
    h7.markdown("**변경 단가**")
    h8.markdown(f"**순단가**<br><small style='color:#888;font-weight:normal'>택배비 {delivery_fee:,}원 제외</small>",
                unsafe_allow_html=True)
    st.markdown("---")

    # 옵션 행 렌더링
    for opt in product.options:
        key = get_weight_key(opt.option_id)
        if key not in st.session_state:
            naver_actual_init = naver_discounted + opt.option_price
            st.session_state[key] = suggest_weight(naver_actual_init, naver_discounted)

        c0, c1, c2, c2b, c3, c4, c5, c6, c7, c8 = st.columns(
            [2.5, 0.7, 0.9, 1.0, 0.8, 1.0, 1.0, 1.0, 1.0, 1.0]
        )

        # ── 고정 기준값 (사용자 기준가와 무관)
        existing_actual = naver_discounted + opt.option_price  # 네이버 실판매가 (고정)
        unit_label, orig_unit_price = parse_unit_price(opt.option_name, existing_actual)

        # 옵션명
        with c0:
            st.write(opt.option_name)

        # 재고
        with c1:
            st.write(f"{opt.stock:,}개")

        # 기존 옵션가 (네이버 상대값)
        with c2:
            if opt.option_price == 0:
                st.write("—")
            elif opt.option_price > 0:
                st.write(f"+{opt.option_price:,}원")
            else:
                st.write(f"{opt.option_price:,}원")

        # 기존 실판매가 (네이버 할인가 + 옵션가 — 고정)
        with c2b:
            st.markdown(f"<b style='color:#1a5276'>{existing_actual:,}원</b>",
                        unsafe_allow_html=True)

        # 가중치 입력
        with c3:
            w = st.number_input(
                "가중치",
                min_value=0.01,
                max_value=20.0,
                value=float(st.session_state[key]),
                step=0.01,
                format="%.2f",
                key=key,
                label_visibility="collapsed",
            )

        # 변경 계산가 = 가중치 × 사용자 기준가
        calc = round(base_price * w)
        with c4:
            st.markdown(f"**{calc:,}원**")

        # 기준가 대비 (변경 계산가 - 사용자 기준가)
        vs_base = calc - base_price
        vs_color = "#1a7f37" if vs_base >= 0 else "#d73a49"
        vs_str = f"+{vs_base:,}" if vs_base >= 0 else f"{vs_base:,}"
        with c5:
            st.markdown(
                f"<span style='color:{vs_color};font-weight:bold'>{vs_str}원</span>",
                unsafe_allow_html=True,
            )

        # 기존 단가 (기존 실판매가 ÷ 팩수 — 고정)
        with c6:
            if orig_unit_price is not None:
                st.markdown(
                    f"<span style='color:#555'>{orig_unit_price:,}원</span>"
                    f"<br><small style='color:#aaa'>{unit_label}</small>",
                    unsafe_allow_html=True,
                )
            else:
                st.write("—")

        # 변경 단가 (변경 계산가 ÷ 팩수)
        _, new_unit_price = parse_unit_price(opt.option_name, calc)
        with c7:
            if new_unit_price is not None:
                diff_unit = new_unit_price - (orig_unit_price or new_unit_price)
                diff_color = "#1a7f37" if diff_unit >= 0 else "#d73a49"
                diff_str = f"+{diff_unit:,}" if diff_unit >= 0 else f"{diff_unit:,}"
                st.markdown(
                    f"<b>{new_unit_price:,}원</b>"
                    f"<br><small style='color:{diff_color}'>{diff_str}</small>",
                    unsafe_allow_html=True,
                )
            else:
                st.write("—")

        # 순단가 (변경 계산가 - 택배비) ÷ 팩수
        net_calc = max(0, calc - delivery_fee)
        _, net_unit_price = parse_unit_price(opt.option_name, net_calc)
        with c8:
            if net_unit_price is not None:
                diff_net = net_unit_price - (new_unit_price or net_unit_price)
                diff_color = "#d73a49" if diff_net < 0 else "#888"
                diff_str = f"{diff_net:,}" if diff_net >= 0 else f"{diff_net:,}"
                st.markdown(
                    f"<b style='color:#1a5276'>{net_unit_price:,}원</b>"
                    f"<br><small style='color:{diff_color}'>{diff_str}</small>",
                    unsafe_allow_html=True,
                )
            else:
                st.write("—")

    st.divider()

    # ──────────────────────────────────────────────
    # ⑥ 계산 결과 요약 테이블
    # ──────────────────────────────────────────────
    st.subheader("⑥ 계산 결과 요약")

    rows = []
    calculated_options = []
    for opt in product.options:
        key = get_weight_key(opt.option_id)
        w = float(st.session_state.get(key, 1.0))
        calc_price = round(base_price * w)
        existing_actual = naver_discounted + opt.option_price   # 고정된 네이버 실판매가
        vs_base = calc_price - base_price
        unit_label, unit_price = parse_unit_price(opt.option_name, calc_price)

        _, orig_unit_price = parse_unit_price(opt.option_name, existing_actual)
        net_calc_price = max(0, calc_price - delivery_fee)
        _, net_unit_price = parse_unit_price(opt.option_name, net_calc_price)

        rows.append({
            "옵션명": opt.option_name,
            "기존 실판매가": existing_actual,
            "가중치": w,
            "변경 계산가": calc_price,
            "기준가 대비": vs_base,
            "기존 단가": f"{orig_unit_price:,}원 ({unit_label})" if orig_unit_price else "—",
            "변경 단가": f"{unit_price:,}원 ({unit_label})" if unit_price else "—",
            f"순단가(택배-{delivery_fee:,})": f"{net_unit_price:,}원 ({unit_label})" if net_unit_price else "—",
            "재고(개)": opt.stock,
        })

        calculated_options.append(CalculatedOption(
            option_id=opt.option_id,
            option_name=opt.option_name,
            original_price=existing_actual,
            calculated_price=calc_price,
            stock=opt.stock,
            weight=str(w),
        ))

    df = pd.DataFrame(rows)
    avg_before = df["기존 실판매가"].mean()
    avg_after = df["변경 계산가"].mean()

    # 통계 배너
    s1, s2, s3, s4, s5 = st.columns(5)
    s1.metric("기준가", f"{base_price:,}원")
    s2.metric("변경 판매가", f"{new_sale_price:,}원")
    s3.metric("평균 기존 실판매가", f"{avg_before:,.0f}원")
    s4.metric("평균 변경가", f"{avg_after:,.0f}원")
    s5.metric("평균 변동", f"{avg_after - avg_before:+,.0f}원")

    # 결과 테이블
    def color_vs(v):
        if isinstance(v, (int, float)):
            if v > 0: return "color: #1a7f37; font-weight:bold"
            if v < 0: return "color: #d73a49; font-weight:bold"
        return ""

    styled = (
        df.style
        .format({
            "기존 실판매가": "{:,}원",
            "가중치": "{:.2f}",
            "변경 계산가": "{:,}원",
            "기준가 대비": lambda v: f"+{v:,}원" if v >= 0 else f"{v:,}원",
            "재고(개)": "{:,}",
        })
        .map(color_vs, subset=["기준가 대비"])
    )
    st.dataframe(styled, use_container_width=True, hide_index=True, height=340)

    # ──────────────────────────────────────────────
    # ⑦ 엑셀 다운로드 (⑤ 옵션별 가중치 설정 테이블 기준)
    # ──────────────────────────────────────────────
    st.divider()
    st.subheader("⑦ 엑셀 다운로드")

    # ⑤ 테이블과 동일한 행 데이터 구성
    excel_option_rows = []
    for opt in product.options:
        key = get_weight_key(opt.option_id)
        w = float(st.session_state.get(key, 1.0))
        existing_actual_xl = naver_discounted + opt.option_price
        calc_price_xl = round(base_price * w)
        vs_base_xl = calc_price_xl - base_price
        unit_label_xl, orig_up = parse_unit_price(opt.option_name, existing_actual_xl)
        _, new_up = parse_unit_price(opt.option_name, calc_price_xl)
        net_calc_xl = max(0, calc_price_xl - delivery_fee)
        _, net_up = parse_unit_price(opt.option_name, net_calc_xl)

        def fmt_opt_price(p):
            if p == 0: return "—"
            return f"+{p:,}원" if p > 0 else f"{p:,}원"

        def fmt_unit(up, label):
            return f"{up:,}원/{label}" if up else "—"

        excel_option_rows.append({
            "옵션명":     opt.option_name,
            "재고":       f"{opt.stock:,}개",
            "기존옵션가": fmt_opt_price(opt.option_price),
            "기존실판매가": f"{existing_actual_xl:,}원",
            "가중치":     round(w, 2),
            "변경계산가": f"{calc_price_xl:,}원",
            "기준가대비": vs_base_xl,
            "기존단가":   fmt_unit(orig_up, unit_label_xl),
            "변경단가":   fmt_unit(new_up, unit_label_xl),
            "순단가":     fmt_unit(net_up, unit_label_xl),
        })

    c_info, c_btn = st.columns([3, 1])
    with c_info:
        st.caption(
            f"파일명: `option_price_{display_id}.xlsx` "
            f"| {len(excel_option_rows)}개 옵션 "
            f"| 기준가 {base_price:,}원 "
            f"| 변경판매가 {new_sale_price:,}원 ({discount_rate}% 할인)"
        )
    with c_btn:
        excel_bytes = build_option_excel(
            product_name=product.product_name,
            product_id=display_id,
            sale_price=product.sale_price,
            discount_amount=product.discount_amount,
            naver_discounted=naver_discounted,
            base_price=base_price,
            new_sale_price=new_sale_price,
            discount_rate=discount_rate,
            delivery_fee=delivery_fee,
            option_rows=excel_option_rows,
        )
        st.download_button(
            label="📥 엑셀 다운로드",
            data=excel_bytes,
            file_name=f"option_price_{display_id}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
