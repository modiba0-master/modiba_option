"""
엑셀 출력 모듈

계산 결과를 openpyxl로 엑셀 파일에 작성합니다.
DB 저장·외부 전송은 일절 수행하지 않습니다.
"""

import io
import logging
from typing import List

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.models.product_model import CalculatedOption, CalculationResult

logger = logging.getLogger(__name__)

_HEADERS = [
    "상품ID",
    "상품명",
    "판매가",
    "기본할인",
    "할인가",
    "옵션ID",
    "옵션명",
    "옵션가 (기존)",
    "변경옵션가 (계산값)",
    "재고수량",
]

_HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
_HEADER_FONT = Font(color="FFFFFF", bold=True, name="맑은 고딕", size=10)
_EVEN_FILL = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
_CENTER = Alignment(horizontal="center", vertical="center")
_LEFT = Alignment(horizontal="left", vertical="center")

_THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def build_excel_bytes(result: CalculationResult) -> bytes:
    """
    CalculationResult를 받아 엑셀 파일을 메모리 상에 생성하고 bytes를 반환합니다.

    Args:
        result: 계산 완료된 상품 결과

    Returns:
        엑셀 파일 bytes (다운로드용)
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "옵션가격계산"

    _write_header(ws)
    _write_rows(ws, result)
    _adjust_column_width(ws)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()


def _write_header(ws) -> None:
    for col_idx, header in enumerate(_HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _CENTER
        cell.border = _THIN_BORDER
    ws.row_dimensions[1].height = 20


def _write_rows(ws, result: CalculationResult) -> None:
    for row_idx, opt in enumerate(result.options, start=2):
        fill = _EVEN_FILL if row_idx % 2 == 0 else None

        values = [
            result.product_id,
            result.product_name,
            result.sale_price,
            result.discount_amount,
            result.discounted_price,
            opt.option_id,
            opt.option_name,
            opt.original_price,
            opt.calculated_price,
            opt.stock,
        ]

        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if fill:
                cell.fill = fill
            cell.border = _THIN_BORDER
            cell.alignment = _CENTER if isinstance(value, (int, float)) else _LEFT


def _adjust_column_width(ws) -> None:
    col_widths = [14, 30, 10, 10, 10, 14, 30, 14, 16, 10]
    for col_idx, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width
